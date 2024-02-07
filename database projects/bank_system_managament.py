import logging
import sqlite3
import time
from sqlite3 import DatabaseError
import bcrypt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta

logging.basicConfig(
    level=logging.DEBUG,  # Set the minimum logging level
    format="%(asctime)s - %(levelname)s - %(message)s"  # Define the format of log messages
)


def get_current_date():
    # this returns the date to other functions
    current_date = datetime.now().date()
    return current_date


def add_days_to_date(date_string, days_to_add):
    # Parse the date string to get year, month, and day
    year, month, day = map(int, date_string.split('/'))

    # Create a datetime object for the input date
    input_date = datetime(year, month, day)

    # Add the specified number of days to the input date
    result_date = input_date + timedelta(days=days_to_add)

    # Format the result date as a string in the format 'yy/mm/dd'
    return result_date.strftime('%y/%m/%d')


def ac_availability(ann):
    temp = False
    try:
        c2 = sqlite3.connect("bank_manage.db")
        try:
            cc5 = c2.cursor()
            q6 = "SELECT account_number FROM personal_bank_account"
            cc5.execute(q6)
            r9 = cc5.fetchall()
            for row in r9:
                if row[0] == ann:
                    temp = True
                    break
            return temp
        except sqlite3.Error as e:
            logging.info("Error while getting details", e)
    except:
        logging.warning("Error while connecting to database")
    finally:
        if c2 in locals():
            c2.close()


def download_sta():
    an7 = int(input("Enter account number: "))
    if ac_availability(an7):
        conn = sqlite3.connect('bank_manage.db')
        query = "SELECT * FROM cash_transaction WHERE ac_no=?"
        df = pd.read_sql_query(query, conn, params=(an7,))
        # Close the database connection
        conn.close()
        wb = Workbook()
        ws = wb.active
        # Write DataFrame to the worksheet starting from cell A1
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        # Apply formatting to the header row
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Apply center alignment to all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        # Save the workbook to a new Excel file
        wb.save('account_statement.xlsx')
        # Log a message indicating that the Excel file is downloaded
        logging.info("account statement is downloaded")
    else:
        print("Account number", an7, "is not found please try again!...")


def save_transaction(account=None, transaction_t=None, dt7=None, a1=None, cih=None, vn=None, fan=None, tan=None):
    com1 = sqlite3.connect("bank_manage.db")
    com1.execute("""INSERT INTO cash_transaction(ac_no, transaction_type, date, amt, cash_in_hand_previous,voucher_no,frm_ac_no,to_ac_no) 
                                                 VALUES (?, ?, ?, ?, ?,?,?,?)""",
                 (account, transaction_t, dt7, a1, cih, vn, fan, tan))
    com1.commit()


def revoke_transaction():
    global res1

    def save_detail(an0, dt, amt1, tt, fan=None, tan=None, vn9=None):
        try:
            nc = sqlite3.connect('bank_manage.db')
            try:
                q4 = ("INSERT INTO revoke_history(ac_no,date,amt,tr_ty,from_ac_no,to_ac_no,voucher_no) VALUES (?,?,?,"
                      "?,?,?,?)")
                nc.execute(q4, (an0, dt, amt1, tt, fan, tan, vn9))
                nc.commit()
            except sqlite3.Error as e:
                logging.warning("Error while saving transaction detail: %s", e)
        except sqlite3.Error as e:
            logging.warning("Error while connecting to database: %s", e)
        finally:
            if 'nc' in locals():
                nc.close()

    print("TO revoke a transaction please authenticate with your account")
    un1 = input("enter your username: ")
    pw = input("enter your password: ")
    res = authenticate_user(un1, pw)

    if res:
        ip3 = int(input("Enter voucher no to Revoke: "))
        try:
            connection = sqlite3.connect("bank_manage.db")
            try:
                curr = connection.cursor()
                cq = "SELECT * FROM cash_transaction WHERE voucher_no=?"
                curr.execute(cq, (ip3,))
                res0 = curr.fetchone()

                try:
                    cn = sqlite3.connect("bank_manage.db")
                    try:
                        cur00 = cn.cursor()
                        q9 = "SELECT voucher_no FROM revoke_history WHERE ac_no=?"
                        cur00.execute(q9, (res0[0],))
                        res1 = cur00.fetchone()
                    except sqlite3.Error as e:
                        logging.warning("Error while getting info: %s", e)
                except sqlite3.OperationalError as e:
                    logging.critical("Error while connecting to database: %s", e)
                if res1[0] == ip3:
                    logging.info("The Transaction is already Revoked")
                else:
                    if res0[1] == "CASH DEPOSIT":
                        q = "UPDATE personal_bank_account SET balance=balance - ? WHERE account_number=?"
                        curr.execute(q, (res0[3], res0[0],))
                        connection.commit()
                        date = get_current_date()
                        save_detail(res0[0], date, res0[3], res0[1], vn9=ip3)
                        logging.info("Successfully Transaction is revoked")
                    elif res0[1] == "CASH WITHDRAW":
                        q = "UPDATE personal_bank_account SET balance=balance + ? WHERE account_number=?"
                        curr.execute(q, (res0[3], res0[0],))
                        dt = get_current_date()
                        save_detail(res0[0], dt, res0[3], res0[1], vn9=ip3)
                        connection.commit()
                        logging.info("Successfully Transaction is revoked")
                    elif res0[1] == "TRF":
                        logging.info("Successfully Transaction is revoked")
                        pass
                    else:
                        print("Unexpected error")
            except sqlite3.OperationalError as e:
                logging.warning("Error while getting transaction details: %s", e)
        except sqlite3.Error as e:
            logging.critical("Error while connecting to database: %s", e)
        finally:
            if 'connection' in locals():
                connection.close()


def get_voucher_no():
    try:
        conn = sqlite3.connect("bank_manage.db")
        cursor = conn.cursor()
        vn = cursor.execute("SELECT voucher_no FROM other")
        vn1 = vn.fetchone()
        conn.close()
        return vn1[0] if vn1 else None
    except sqlite3.Error as e:
        logging.error("Error fetching voucher number: %s", e)
        return None


def create_ac():
    # it creates the account no in the database by using simple query INSERT
    nm = input("Enter The Customer Full Name")
    at = input("Enter The Account Type")
    ad = input("enter Customer Full Address")
    mn = input("Enter The Customer Mobile NO.")
    an = input("Enter the Customer Aadhar NO.")
    pn = input("Enter The Pan card NO")
    con = sqlite3.connect("bank_manage.db")
    dt = get_current_date()
    con.execute("""
        INSERT INTO personal_bank_account
        (account_holder_name, balance, account_type, opening_date, address, mobile_number, aadhar_card_number, pan_card_number)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, (nm, 0, at, dt, an, pn, an, pn))
    con.commit()
    con.close()
    print("...........Customer Account Created successfully........")


class Customer:
    # This class is separately for Editing the customer personal information only
    # it gives parameter by using its constructor is only a account no.
    def __init__(self, an0):
        self.an = an0

    def change_pan(self):
        pn = input("Enter pan card Number to Update")
        cono = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET pan_card_number=? WHERE account_number=?"
        cono.execute(q, (pn, self.an))
        cono.commit()
        cono.close()
        logging.info("Pan Card Number Updated Successfully")

    def change_card_number(self):
        pn = input("Enter aadhar card Number to Update")
        con = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET aadhar_card_number=? WHERE account_number=?"
        con.execute(q, (pn, self.an))
        con.commit()
        con.close()

        logging.info("Pan Card Number Updated Successfully")

    def change_mobile_no(self):
        mn = input("Enter Mobile card Number to Update")
        con = sqlite3.connect("bank_manage.db")
        q = "UPDATE personal_bank_account SET mobile_number=? WHERE account_number=?"
        con.execute(q, (mn, self.an))
        con.commit()
        logging.info("Mobile Number Updated Successfully")

    def get_balance(self):
        try:
            con = sqlite3.connect("bank_manage.db")
            cur = con.cursor()
            q = "SELECT balance FROM personal_bank_account WHERE account_number=?"
            cur.execute(q, (self.an,))
            result2 = cur.fetchone()
            con.close()  # Move connection closure after fetching the result
            return result2[0] if result2 else None
        except sqlite3.Error as e:
            logging.error("Error fetching balance: %s", e)
            return None


class Wallet:
    # This class is for only purpose of handling cash in hand transaction
    # this class get initial amt amount that we give by fetching from our database at the time of calling
    # Firstly its return the cash in hand or Another function set the cash in hand
    def __init__(self, set_amt):
        self.set_amt = set_amt

    def cash_in_hand_return(self):
        return self.set_amt

    def cash_in_hand_deposit(self):
        try:
            con3 = sqlite3.connect("bank_manage.db")
            try:
                q = "UPDATE cash_in_hand SET cash_in_hand=cash_in_hand + ?"
                con3.execute(q, (self.set_amt,))
                con3.commit()
                con3.close()
            except sqlite3.Error as e:
                con3.rollback()
                logging.info("Transaction Failed.....", e)
        except sqlite3.OperationalError:
            logging.warning("Error while connecting to database")

    def cash_in_hand_withdraw(self):
        try:
            con3 = sqlite3.connect("bank_manage.db")
            try:
                q = "UPDATE cash_in_hand SET cash_in_hand=cash_in_hand - ?"
                con3.execute(q, (self.set_amt,))
                con3.commit()
                con3.close()
            except sqlite3.Error as e:
                con3.rollback()
                logging.info("Transaction was failed...", e)
        except sqlite3.Error as error:
            logging.warning("Error while connecting to Database...")


class Account:
    # This is major class of this program it's having some main member functions are included in it this class
    # inherits the class Wallet in it Class Wallet have an one argument called initial Class Account have one
    # argument called an Account number Because of this argument or Inheritance we have to call a base class
    # constructor manually that why pythons allowed to call a manual constructor
    def __init__(self, an2, amt6):
        self.amt2 = amt6
        self.an2 = an2

    def deposit(self):
        # Here we have to deposit cash in customers balance so we have to do firstly give an ac no. to class then give
        # amt to the member function and the class wallet get its their initialized value from the fetching if else by
        # fetching from database so now only we have to do adding and subtraction logic in the member function for
        # respective
        try:
            con4 = sqlite3.connect("bank_manage.db")
            try:
                q1 = "UPDATE personal_bank_account SET balance=balance + ? WHERE account_number = ?"
                Q2 = "UPDATE other SET voucher_no=voucher_no+1"
                con4.execute(q1, (self.amt2, self.an2))
                con4.execute(Q2)
                con4.commit()
                ob = Wallet(self.amt2)
                vn = int(get_voucher_no())
                ob.cash_in_hand_deposit()
                if vn is not None:
                    logging.info("Amount deposited to AC NO.%s voucher no is.%s", self.an2, vn + 1)
                else:
                    logging.warning("Error while getting to fetch voucher no.")
                ob = Customer(self.an2)
                ab = ob.get_balance()
                dt = get_current_date()
                tt = "CASH DEPOSIT"
                save_transaction(self.an2, tt, dt, self.amt2, ab - self.amt2, vn)  # Calling the function
            except sqlite3.Error as e:
                if con4 is not None:
                    con4.rollback()
                logging.info("Transaction failed:", e)
        except sqlite3.Error as e:
            logging.warning("Database connection failed....", e)
        finally:
            if 'con4' and 'com1' in locals():
                con4.close()

    def withdraw(self):
        try:
            con6 = sqlite3.connect("bank_manage.db")
            try:
                q1 = "UPDATE personal_bank_account SET balance=balance - ? WHERE account_number = ?"
                Q2 = "UPDATE other SET voucher_no=voucher_no+1"
                con6.execute(q1, (self.amt2, self.an2))
                con6.execute(Q2)
                con6.commit()
                ob = Wallet(self.amt2)
                ob.cash_in_hand_withdraw()
                vn = int(get_voucher_no())
                if vn is not None:
                    logging.info("Amount Withdrawn from AC NO.%s and voucher no is.%s", self.an2, vn + 1)
                else:
                    logging.warning("Error while getting voucher no.")
                ob = Customer(self.an2)
                ab = ob.get_balance()
                dt = get_current_date()
                tt = "CASH WITHDRAW"
                save_transaction(self.an2, tt, dt, self.amt2, ab - self.amt2, vn)  # Calling the function
            except sqlite3.Error as e:
                if con6 is not None:
                    con6.rollback()
                logging.info("Transaction Failed", e)
        except sqlite3.OperationalError as e:
            logging.info("Error while connecting to database...", e)
        finally:
            if 'con6' in locals():
                con6.close()

    def change_balance(self, amt2):
        pass

    def transfer(self, a02):
        global rr, cc, cc2
        try:
            cc = sqlite3.connect("bank_manage.db")
            try:
                cc2 = cc.cursor()
                qq = "SELECT balance FROM personal_bank_account WHERE account_number=?"
                cc2.execute(qq, (self.an2,))
                rr = cc2.fetchone()
            except sqlite3.Error as e:
                logging.warning("Error while getting details...", e)
        except:
            logging.warning("Error while connecting to database...", exc_info=True)
        if rr[0] > self.amt2:
            qq1 = "UPDATE personal_bank_account SET balance=balance-? WHERE account_number=?"
            qq2 = "UPDATE personal_bank_account SET balance=balance+? WHERE account_number=?"
            try:
                cc2.execute(qq1, (self.amt2, self.an2))
                try:
                    cc2.execute(qq2, (self.amt2, a02))
                    cc.commit()
                    tt = "TRF"
                    d = get_current_date()
                    vn = get_voucher_no()
                    save_transaction(transaction_t=tt, dt7=d, vn=vn, fan=self.an2, tan=a02)  # Calling the function
                    logging.info("%s RS. successfully Transferred to account number: %s", self.amt2, a02)
                except sqlite3.Error as e:
                    cc.rollback()
                    logging.warning("Transfer failed due to", e)
            except:
                logging.warning("Transaction failed....", exc_info=True)
            finally:
                cc.close()
        else:
            logging.info("The Account Balance Is Too Low To Debit The Amt")

    def apply_interest(self, intrest, period):
        pass

    def close_ac(self):
        an7 = int(input("Enter account number to close"))
        try:
            cc3 = sqlite3.connect('bank_manage.db')
            try:
                cc0 = cc3.cursor()
                q = "DELETE FROM personal_bank_account WHERE ac_no=?"
                qq2 = "SELECT * FROM personal_bank_account WHERE ac_no=?"
                result0 = cc0.execute(qq2, (an7,))
                cc3.commit()
                try:
                    cc0.execute(q, (an7,))
                    cc3.commit()
                except:
                    logging.info("Error while getting to save details")
            except sqlite3.Error as e:
                logging.info("Error while trying to get details")
        except:
            logging.info("Error while trying to connect to database")
        finally:
            cc3.close()


class Transaction:
    pass


class Common_functions:
    def __init__(self, AMT, cr_ac, debit_ac):
        self.AMT = AMT
        self.cr_ac = cr_ac
        self.debit_ac = debit_ac

    def transfer_funds(self):
        try:
            # Connect to the SQLite database
            connection = sqlite3.connect("bank_manage.db")
            cursor = connection.cursor()
            # Begin transaction
            cursor.execute("BEGIN TRANSACTION")
            # Deduct amount from debit account
            cursor.execute("UPDATE personal_bank_account SET balance = balance - ? WHERE account_number = ?",
                           (self.AMT, self.debit_ac))
            # Add amount to credit account
            cursor.execute("UPDATE fd_accounts SET account_balance = ? WHERE fd_ac_no = ?",
                           (self.AMT, self.cr_ac))
            connection.commit()
            # save to the sql table
            save_transaction(transaction_t="TRF", dt7=get_current_date(), vn=get_voucher_no(), fan=self.debit_ac,
                             tan=self.cr_ac)
        except sqlite3.DatabaseError as e:
            logging.error(f"Database error: {e}")
            # Rollback transaction if an error occurs
            connection.rollback()
        except Exception as e:
            logging.error(f"Error: {e}")
            # Rollback transaction if an error occurs
            connection.rollback()
        finally:
            # Close the database connection
            if connection:
                connection.close()

    def self_bank_transfer(self):
        pass


class Fdaccount:
    def __init__(self, an88=None):
        self.an88 = an88

    def check_account(self, ann):
        temp = False
        try:
            c2 = sqlite3.connect("bank_manage.db")
            try:
                cc5 = c2.cursor()
                q6 = "SELECT fd_ac_no FROM fd_accounts"
                cc5.execute(q6)
                r9 = cc5.fetchall()
                print(r9)
                for row in r9:
                    if row[0] == ann:
                        temp = True
                        break
                return temp
            except sqlite3.Error as e:
                logging.info("Error while getting details", e)
        except:
            logging.warning("Error while connecting to database")
        finally:
            if c2 in locals():
                c2.close()

    def get_account_number(self):
        try:
            cc30 = sqlite3.connect('bank_manage.db')
            try:
                cc30c = cc30.cursor()
                q01 = "SELECT fd_ac_no FROM fd_accounts"
                cc30c.execute(q01)
                r9 = cc30c.fetchall()
                return r9[-1]
            except Exception as e:
                logging.info("Error getting to get account number", e)
        except sqlite3.Error as e:
            logging.info("Error while connecting to database")

    def calculate_mat_amt(self, fd_amt, ir1, fd_days1):
        interest1 = (fd_amt * ir1 * fd_days1) / (365 * 100)
        return interest1

    def days_between_dates(self, date1_str, date2_str):
        # Convert the date strings to datetime objects
        date1 = datetime.strptime(date1_str, '%Y-%m-%d')
        date2 = datetime.strptime(date2_str, '%Y-%m-%d')
        delta = date2 - date1
        # Return the absolute value of the number of days
        return abs(delta.days)

    def create_account(self):
        print("Enter the saving account number to open an FD account")
        pan = int(input())
        if ac_availability(pan):
            try:
                cc1 = sqlite3.connect('bank_manage.db')
                cursor1 = cc1.cursor()
                q01 = "SELECT * FROM personal_bank_account WHERE account_number=?"
                cursor1.execute(q01, (pan,))
                rro = cursor1.fetchone()
                if rro:
                    fd_am = int(input("Enter Amount to make an FD: "))
                    ir = float(input("Enter interest rate to give to customer: "))
                    fd_days = int(input("Enter number of days to make an FD: "))
                    od = str(get_current_date())
                    cn50 = rro[1]

                    try:
                        qq2 = ("INSERT INTO fd_accounts (customer_name, opening_date, int_rate, fd_days, "
                               "pre_mature_withdraw,fd_opening_amt, mat_amt,personal_ac_no,fd_mat_dt) VALUES (?, ?, ?, ?, ?, ?,"
                               "?,?,?)")
                        mat_amt = self.calculate_mat_amt(fd_am, ir, fd_days)
                        cursor1.execute(qq2, (
                            rro[1], od, ir, fd_days, False, fd_am, mat_amt + fd_am, pan, add_days_to_date(od, fd_days)))
                        cc1.commit()
                        fac = self.get_account_number()
                        logging.info("Account created successfully With account number %s",
                                     fac[0])
                    except sqlite3.Error as e:
                        logging.error("Account creation failed: %s", e)
                else:
                    logging.warning("No such account number found: %d", pan)
            except sqlite3.Error as e:
                logging.error("Error while trying to connect to database: %s", e)
            finally:
                cc1.close()
        else:
            print("No such account number is found. Please try again.")

    def delete_account(self):
        pass

    def check_fd_balance(self, fd_ac_n):
        # check FD balance and FD opening balance and returns tuple
        try:
            o = sqlite3.connect("bank_manage.db")
            try:
                c1 = o.cursor()
                qq2 = "SELECT account_balance,fd_opening_amt FROM fd_accounts WHERE fd_ac_no=?"
                c1.execute(qq2, (fd_ac_n,))
                r22 = c1.fetchone()
                return r22
            except sqlite3.Error as r:
                logging.info("Error while getting details", r)
        except:
            logging.error("Error while connecting to database")

    def add_funds(self):
        # get Account number
        an5 = int(input("Enter FD Account number to add funds:"))
        # Check account availability
        if self.check_account(an5):
            # CHECK FD BALANCE FOR AVOIDING REPEAT TRANSACTION
            r3 = self.check_fd_balance(an5)
            if r3[0] == r3[1]:
                logging.info("FD Amount is Already Transferred....")
            else:
                print("1.Transfer from the Saving Account")
                print("2.For Cash deposit")
                print("3.Online Transfer")
                c = int(input("Enter your choice"))
                if c == 1:
                    try:
                        c = sqlite3.connect("bank_manage.db")
                        try:
                            # Getting all details form fd_ac_table information for specific account number
                            cu = c.cursor()
                            query8 = "SELECT * FROM fd_accounts WHERE fd_ac_no=?"
                            cu.execute(query8, (an5,))
                            rows = cu.fetchone()
                            cc1 = Customer(rows[9])
                            result01 = cc1.get_balance()
                            # comparison between saving account balance with fd balance to Transfer
                            if result01 > rows[8]:
                                ob_co = Common_functions(rows[8], rows[0], rows[9])
                                try:
                                    ob_co.transfer_funds()
                                    logging.info("AMT Transfer successfully from Saving Account To FD Account....")
                                except Exception as e:
                                    logging.error("Transfer Failed due to", e)
                            else:
                                logging.info("The Personal Account Balance Is To Low For Add funds in FD Account")
                        except sqlite3.Error as err:
                            logging.info("Error Occurred", err)
                            c.rollback()
                    except Exception as e:
                        logging.info("Error while connecting to database", e)
                    finally:
                        c.close()
                elif c == 2:
                    try:
                        conec = sqlite3.connect("bank_manage.db")
                        ccu = conec.cursor()
                        qq8 = "SELECT account_balance,fd_opening_amt FROM fd_accounts WHERE fd_ac_no=?"
                        ccu.execute(qq8, (an5,))
                        tr = ccu.fetchone()
                        rr = self.check_fd_balance(an5)
                        if rr[0] == tr[0]:
                            logging.info("AMT Is Already Transferred")
                        else:
                            print("Enter Deposit Amount For FD Number.", an5)
                            cda = int(input())
                            if tr[1] == cda:
                                try:
                                    qq1 = "UPDATE fd_accounts SET account_balance WHERE fd_ac_no=?"
                                    ccu.execute(qq1, (an5,))
                                    conec.commit()
                                    try:
                                        ow = Wallet(cda)
                                        ow.cash_in_hand_deposit()
                                    except Exception as e:
                                        logging.error(e)
                                        conec.rollback()
                                except Exception as e:
                                    conec.rollback()
                                    logging.info("Transaction is failed")
                                finally:
                                    conec.close()
                            else:
                                print("Entered Amount is wrong to Deposit")
                    except DatabaseError as e:
                        logging.error("Error while connecting to database", e)
                elif c == 3:
                    pass
                else:
                    print("Invalid Input")
        else:
            logging.info("Account does not exist. Please try again.")

    def withdraw(self):
        fd_Ac_n = int(input("Enter the Fd Account Number: "))
        if self.check_account(fd_Ac_n):
            try:
                ccu = sqlite3.connect("bank_manage.db")
                ccu1 = ccu.cursor()
                qq8 = "SELECT * FROM fd_accounts WHERE fd_ac_no=?"
                ccu1.execute(qq8, (fd_Ac_n,))
                fd_res1 = ccu1.fetchone()
                dt = get_current_date()
                if dt == fd_res1[10]:
                    print("Maturity Date is Over you can proceed to Withdraw FD")
                    if fd_res1[2] == fd_res1[8]:
                        try:
                            q11 = "UPDATE fd_accounts SET account_balance=account_balance - ? WHERE fd_ac_no=?"
                            ccu1.execute(q11, (fd_res1[2], fd_Ac_n))
                            ccu.commit()
                            try:
                                q12 = "UPDATE personal_bank_account SET balance=balance + ? WHERE account_number=?"
                                ccu1.execute(q12, (fd_res1[7], fd_res1[9]))
                                ccu.commit()
                                logging.info("Successfully Amount Transferred to Saving account")
                            except Exception as e:
                                ccu.rollback()
                                logging.error("Transfer Failed", e)
                        except Exception as e:
                            ccu.rollback()
                            logging.error("Transfer failed to personal Account")
                else:
                    c = int(input("Today is not maturity date Please Confirm To Premature Withdraw Press 1"))
                    if c == 1:
                        try:
                            ccu0 = sqlite3.connect("bank_manage.db")
                            try:
                                ccu10 = ccu0.cursor()
                                cc20 = "SELECT * FROM fd_accounts WHERE fd_ac_no=?"
                                ccu10.execute(cc20, (fd_Ac_n,))
                                r45 = ccu10.fetchone()
                                dd=str(get_current_date())
                                dy = int(self.days_between_dates(dd,r45[3]))
                                try:
                                    qq11 = "UPDATE personal_bank_account SET balance=balance + ? WHERE account_number=?"
                                    ccu1.execute(qq11, (self.calculate_mat_amt(r45[2], r45[4], dy), r45[10]))
                                    ccu0.commit()
                                    try:
                                        qq12 = ("UPDATE fd_accounts SET account_balance=account_balance - ? WHERE "
                                                "fd_ac_no=?")
                                        ccu10.execute(qq12, (r45[2], fd_Ac_n))
                                        ccu0.commit()
                                        logging.info("FD withdraw successfully Amt is Transferred to Saving Account")
                                    except Exception as e:
                                        ccu0.rollback()
                                except DatabaseError as e:
                                    logging.info(e)
                            except sqlite3.Error as e:
                                logging.error("ERROR WHILE FETCHING DATA", e)
                        except sqlite3.Error as e:
                            logging.warning("Database connection failed", e)
                        finally:
                            ccu0.close()
                    else:
                        exit(0)
            except DatabaseError as e:
                logging.error(e)
            finally:
                ccu.close()

        else:
            print("Fd account number is not found")

    def check_maturity(self):
        pass

    def check_interest(self):
        pass

    def extend_date(self):
        pass


class RDaccount:
    def apply_interest(self):
        # Calculate and apply interest to the RD account balance
        pass

    def check_maturity(self):
        # Check if the RD account has matured
        pass

    def withdraw_pre_mature(self):
        # Allow premature withdrawal from the RD account (may not be applicable for RD)
        pass

    def check_interest(self):
        # Check the interest earned on the RD account
        pass

    def check_balance(self):
        # Check the current balance of the RD account
        pass

    def update_account_details(self, monthly_deposit_amount=None, interest_rate=None, duration=None):
        # Update RD account details such as monthly deposit amount, interest rate, or duration
        pass

    def close_account(self):
        # Close the RD account
        pass

    def view_account_statement(self):
        # Provide a statement of account activity
        pass

    def extend_duration(self, additional_duration):
        # Extend the duration of the RD account
        pass


class fd_loan:
    pass


class rd_loan:
    pass


class personal_loan:
    pass


class auto:
    pass


class business_loan:
    pass


class home_laon:
    pass


def authenticate_user(username, entered_password):
    try:
        connection = sqlite3.connect('bank_manage.db')
        cursor = connection.cursor()

        cursor.execute("SELECT hashed_password, bank_name FROM users WHERE username = ?", (username,))
        result = cursor.fetchone()

        connection.close()

        if result:
            stored_hashed_password = result[0]  # Convert hashed password to bytes
            if bcrypt.checkpw(entered_password.encode('utf-8'), stored_hashed_password):
                logging.info("Authentication successful!")
                return True, result[1]
            else:
                logging.warning("Authentication failed.")
                return False
        else:
            logging.warning("User not found.")
            return False
    except TypeError as e:
        logging.critical("An error occurred during authentication: %s", e)
        return None


print(".........BANK MANAGEMENT SOFTWARE.........")
print("PlEASE LOGIN FIRST")
un = input("Enter User name")
pa = input("Enter Password")
result = authenticate_user(un, pa)
flag = True
while True:
    if result:
        time.sleep(0.5)
        if flag:
            print(".................Welcome", result[1], "..............")
            flag = False
        print("Select main menu")
        print("1.Customer Management")
        print("2.Advanced Accounts")
        print("3.Transaction Management")
        print("4.Data Analysis And Reporting")
        print("5.Log Out")
        ch = int(input("Enter You Choice"))
        if ch == 1:
            print("Select following operations")
            print("1.Create Account")
            print("2.View Account Detail")
            print("3.Deposit Money")
            print("4.Withdraw money")
            print("5.Transfer money")
            print("6.Check Balance")
            print("7.Account Statement")
            print("8.Close Account")
            print("9.Edit personal Account detail")
            print("10.Log Out")
            ch1 = int(input("Enter your Choice"))
            if ch1 == 1:
                create_ac()
            elif ch1 == 2:
                pass
            elif ch1 == 3:
                an = int(input("Enter Account No.To Deposit Money"))
                if ac_availability(an):
                    amt = int(input("Enter Amt to Deposit"))
                    obj = Account(an,
                                  amt)  # passed the account number and the amt passed to The class Account constructor
                    obj.deposit()
                else:
                    print("Account Not Available")
            elif ch1 == 4:
                an = int(input("Enter Account No.To withdraw"))
                if ac_availability(an):
                    amt = int(input("Enter Amt to withdraw"))
                    obj = Account(an, amt)
                    obj.withdraw()
                else:
                    print("Account number not found...")
            elif ch1 == 5:
                payee_account = int(input("Please input the account number from which you want to withdraw or debit "
                                          "funds: "))
                if ac_availability(payee_account):
                    transfer_account = int(input("Please enter the recipient's account number for the transfer: "))
                    if ac_availability(transfer_account):
                        transfer_amount = int(input("Please specify the amount you wish to transfer: "))
                        obj3 = Account(payee_account, transfer_account)
                        obj3.transfer(transfer_amount)
                    else:
                        print("Account number,", transfer_account, "Not found...")
                else:
                    print("Account number", payee_account, "Not found")
            elif ch1 == 6:
                an8 = int(input("Enter account number"))
                obj3 = Customer(an8)
                r = obj3.get_balance()
                print("Your Balance is ", r)
            elif ch1 == 7:
                download_sta()
            elif ch1 == 8:
                pass
            elif ch1 == 9:
                ip = int(input("Enter Account No.To Edit"))
                if ac_availability(ip):
                    AC = Customer(ip)
                    print("Enter 1 for edit pan card")
                    print("2 for edit mobile no.")
                    print("3 for edit aadhar")
                    ip2 = int(input())
                    if ip2 == 1:
                        AC.change_pan()
                    elif ip2 == 2:
                        AC.change_mobile_no()
                    elif ip2 == 3:
                        AC.change_card_number()
                else:
                    print("Account number not found")
            elif ch == 10:
                pass
        elif ch == 2:
            print("Select following operation")
            print("1.Fixed Deposit Account")
            print("2.Create Recurring Deposit Account")
            print("3.Grant a Loan")
            ch = int(input("Enter your Choice"))
            if ch == 1:
                print("Enter 1 For Create Account ")
                print("Enter 2 For Add funds To FD Account")
                print("Enter 3 For Withdraw Account ")
                chh = int(input("Enter your Choice"))
                fd_obj = Fdaccount()
                if chh == 1:
                    fd_obj.create_account()
                elif chh == 2:
                    fd_obj.add_funds()
                elif chh == 3:
                    fd_obj.withdraw()
                else:
                    print("Invalid choice")
            elif ch == 2:
                pass
            elif ch == 3:
                pass

        elif ch == 3:
            print("Enter Your Choice")
            print("1.View Day book")
            print("2.Analyze Cash in Hand")
            print("3.Customer Transaction History")
            print("4.Revoke transaction ")
            ch = int(input("Enter Your Choice"))
            if ch == 1:
                pass
            elif ch == 2:
                pass
            elif ch == 3:
                pass
            elif ch == 4:
                revoke_transaction()
        elif ch == 4:
            nn = int(input("Enter ac no"))
            ac_availability(nn)
            print("Enter your choice")
            print("1.Total Accounts")
            print("2.Balance Sheet")
        elif ch == 5:
            logging.info("Thank U for using our Bank System")
            exit(0)
    else:
        break
